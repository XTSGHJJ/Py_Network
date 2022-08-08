from re import search,findall
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side
from os import listdir
from PySimpleGUI import popup_get_folder,popup,popup_ok

def HW_57_RE(txt):
    HW_57_Name = search(r'sysname (.*)',txt)
    HW_57_Sn = search(r'.*(210[A-Z0-9]{17})',txt)
    HW_57_Version = search(r'.*(Version.*)',txt)
    HW_57_Uptime = search(r'.*(uptime.*)',txt)
    HW_57_CPU = search(r'.*(five seconds: \d+%: one minute: \d+%: five minutes: \d+%)',txt)
    HW_57_Memory = search(r'System Total Memory Is: (\d+.*)\n Total Memory Used Is: (\d+.*)\n Memory Using Percentage Is: (\d+%)',txt)
    HW_57_Fan = search(r'\s(\d)\s+(\d)\s+([A-Za-z]+)\s+([A-Za-z]+)\s.*',txt)
    #HW_57_Power = search(r'',txt)
    #HW_57_Log = re.search(r'.*%%\d+\w+/(\d)/.*',txt)
    info_list =[HW_57_Name,HW_57_Sn,HW_57_Version,HW_57_Uptime,HW_57_CPU,HW_57_Memory,HW_57_Fan]
    file_result = []
    for i in info_list:
        if not i == None:
            file_result.append(i.group(1))
        else:
            file_result.append(i)
    return file_result

def HW_68_RE(txt):
    pass

def H3C_57_RE(txt):
    H3C_57_Name = search(r'sysname (.*)',txt)
    H3C_57_SN = search(r'DEVICE_SERIAL_NUMBER : (210[A-Z0-9]{17})',txt)
    H3C_57_Version = search(r'.*(Version.*)',txt)
    H3C_57_Uptime = search(r'.*(uptime.*)',txt)
    H3C_57_CPU = search(r'\s+(\d+)%\sin\slast\s\d\s\w+\n\s+(\d+)%\sin\slast\s\d\s\w+\n\s+(\d+)%\sin\slast\s\d\s\w+',txt)
    H3C_57_Memory = search(r'Mem:.*\s+(\d+\.?\d+)%',txt)
    H3C_57_Fan = search(r' State    : (.*)',txt)
    H3C_57_Power = search(r' \d+\s+(\w+)\s+AC.*',txt)
    info_list =[H3C_57_Name,H3C_57_SN,H3C_57_Version,H3C_57_Uptime,H3C_57_CPU,H3C_57_Memory,H3C_57_Fan,H3C_57_Power]
    file_result = []
    x = 0
    for i in info_list:
        if not i == None:
            if x == 4:
                five_second = i.group(1)
                one_minute = i.group(2)
                five_minute = i.group(3)
                all_cpu_usage =(five_second + '-' + one_minute + '-' + five_minute)
                file_result.append(all_cpu_usage)
            elif x == 5:
                memory_usage = 100 - float(i.group(1))
                file_result.append(str("{:.2f}".format(memory_usage)) + '%')
            else:
                file_result.append(i.group(1))
        else:
            file_result.append(i)
        x += 1
    return file_result

def H3C_68_RE(txt):
    pass

def FG_RE(txt):
    pass

def CISCO_RE(txt):
    pass

def SG_RE(txt):
    pass

def judge(content):
    #Huawei
    if findall(r'Huawei',content):
        if findall(r'S5\d{3}',content):
            return HW_57_RE(content)
        elif findall(r'CE\d.*',content): 
            return HW_68_RE(content)
    #New H3C
    if findall(r'H3C',content):
        if findall(r'S5\d{3}',content):
            return H3C_57_RE(content)
        elif findall(r'S6\d{3}',content):
            return H3C_68_RE(content)
    #Fortinet
    if findall(r'ForiGate',content):
        return  FG_RE(content)
    #Cisco
    if findall(r'Cisco',content):
        return CISCO_RE(content)
    #Hillstone
    if findall(r'StoneOS',content):
        return SG_RE(content)

def main(file_path):
    #初始化第一行表头
    wb = Workbook()
    ws = wb.active
    ws.title = '巡检信息表'
    init_header = ['Device-Name','SN','Version','Uptime','CPU-Usage','Memory-Usager','Fan-Status','Power-Status']
    ws.append(init_header)
    yellowFill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for i in range(1,len(init_header)+1):
        ws.cell(row=1, column=i,).fill = yellowFill
        ws.cell(row=1, column=i,).border = thin_border
    
    for file_name in listdir(file_path): #遍历目标文件夹
        with open (file_path + file_name,'r', encoding='utf-8', errors='ignore') as file:
            search_file = file.read()
            info = judge(search_file) #判断设备类型
            #print(info)
            ws.append(info)
    wb.save('Inspection_Info.xlsx')

if __name__ == '__main__':
    # 窗口显示文本框和浏览按钮, 以便选择一个文件夹
    path = popup_get_folder("Select Folder")
    if not path:
        popup("Cancel", "No folder selected")
        raise SystemExit("Cancelling: no folder selected")
    main(path + '/')
    popup_ok('Successfully Completed!')






